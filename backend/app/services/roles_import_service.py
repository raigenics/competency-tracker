"""
Roles import service for Excel bulk import.

SRP: Handles Excel parsing and duplicate detection for roles import.
Duplicate detection checks bidirectionally:
- role_name vs existing role_names
- role_name vs existing alias tokens
- alias tokens vs existing role_names
- alias tokens vs existing alias tokens
"""
import logging
from typing import List, Dict, Set, Tuple, Optional
from dataclasses import dataclass
from datetime import datetime, timezone
from io import BytesIO
import pandas as pd
from openpyxl import load_workbook
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError

from app.models.role import Role
from app.utils.normalization import normalize_key

logger = logging.getLogger(__name__)


@dataclass
class RoleImportRow:
    """Represents a single row from the roles import Excel file."""
    row_number: int
    role_name: str
    alias: Optional[str]  # Comma-separated alias string
    description: Optional[str]


@dataclass
class ImportResult:
    """Result of the import operation."""
    total_rows: int
    success_count: int
    failure_count: int
    failures: List[Dict]  # {row_number, role_name, reason}
    imported_roles: List[Dict]  # Successfully imported role data


class RoleImportService:
    """Service for importing roles from Excel files.
    
    Performs O(1) duplicate detection using in-memory token maps.
    """
    
    # Expected column names (case-insensitive)
    EXPECTED_COLUMNS = ["role name", "alias", "role description"]
    
    def __init__(self, db: Session):
        self.db = db
        self.errors: List[Dict] = []
        
        # Token maps for O(1) duplicate detection
        self.role_name_map: Dict[str, int] = {}  # normalized name -> role_id
        self.alias_token_map: Dict[str, int] = {}  # normalized token -> role_id
    
    def import_roles(self, file_content: bytes, created_by: str = "system") -> ImportResult:
        """
        Import roles from Excel file.
        
        Args:
            file_content: Bytes content of the Excel file
            created_by: User performing the import
            
        Returns:
            ImportResult with counts and failure details
        """
        # Parse Excel file
        rows = self._parse_excel(file_content)
        
        if not rows:
            return ImportResult(
                total_rows=0,
                success_count=0,
                failure_count=0,
                failures=self.errors,
                imported_roles=[]
            )
        
        # Build conflict detection maps from existing roles
        self._build_conflict_maps()
        
        # Process rows
        failures = []
        imported = []
        
        # Track new names/aliases for intra-batch conflict detection
        batch_names: Set[str] = set()
        batch_aliases: Set[str] = set()
        
        for row in rows:
            # Validate role name
            if not row.role_name or not row.role_name.strip():
                failures.append({
                    "row_number": row.row_number,
                    "role_name": row.role_name or "(empty)",
                    "reason": "Role name is required"
                })
                continue
            
            role_name = row.role_name.strip()
            role_name_norm = normalize_key(role_name)
            
            # Parse alias tokens
            alias_tokens = self._parse_alias_tokens(row.alias)
            alias_tokens_norm = [normalize_key(t) for t in alias_tokens]
            
            # Check for conflicts
            conflict = self._check_conflicts(
                role_name_norm, alias_tokens_norm, batch_names, batch_aliases
            )
            
            if conflict:
                failures.append({
                    "row_number": row.row_number,
                    "role_name": role_name,
                    "reason": conflict
                })
                continue
            
            # Create the role using SAVEPOINT for transaction isolation
            # This ensures a single failing row doesn't break subsequent rows
            try:
                with self.db.begin_nested():
                    new_role = Role(
                        role_name=role_name,
                        role_alias=row.alias.strip() if row.alias else None,
                        role_description=row.description.strip() if row.description else None,
                        created_at=datetime.now(timezone.utc),
                        created_by=created_by
                    )
                    
                    self.db.add(new_role)
                    self.db.flush()  # Get role_id
                
                # Update conflict maps for subsequent rows
                self.role_name_map[role_name_norm] = new_role.role_id
                for token_norm in alias_tokens_norm:
                    if token_norm:
                        self.alias_token_map[token_norm] = new_role.role_id
                
                # Track in batch
                batch_names.add(role_name_norm)
                batch_aliases.update(t for t in alias_tokens_norm if t)
                
                imported.append({
                    "role_id": new_role.role_id,
                    "role_name": new_role.role_name,
                    "role_alias": new_role.role_alias,
                    "role_description": new_role.role_description
                })
                
            except IntegrityError as e:
                # Map DB unique constraint violation to friendly message
                # SAVEPOINT is automatically rolled back by begin_nested() context manager
                logger.warning(f"IntegrityError importing role '{role_name}': {e}")
                failures.append({
                    "row_number": row.row_number,
                    "role_name": role_name,
                    "reason": "Duplicate: Role name already exists"
                })
            except Exception as e:
                # Generic error - don't leak DB details
                # SAVEPOINT is automatically rolled back by begin_nested() context manager
                logger.error(f"Unexpected error importing role '{role_name}': {e}", exc_info=True)
                failures.append({
                    "row_number": row.row_number,
                    "role_name": role_name,
                    "reason": "Unexpected error while importing this row"
                })
        
        # Commit successful imports
        if imported:
            self.db.commit()
        
        return ImportResult(
            total_rows=len(rows),
            success_count=len(imported),
            failure_count=len(failures),
            failures=failures,
            imported_roles=imported
        )
    
    def _parse_excel(self, file_content: bytes) -> List[RoleImportRow]:
        """Parse Excel file and return list of RoleImportRow objects."""
        try:
            # Load workbook with read_only=True to skip data validation parsing
            wb = load_workbook(filename=BytesIO(file_content), read_only=True, data_only=True)
            ws = wb.active
            
            # Read all rows
            rows_data = list(ws.iter_rows(values_only=True))
            wb.close()
            
            if not rows_data:
                self.errors.append({
                    "row_number": 0,
                    "role_name": "",
                    "reason": "Excel file is empty"
                })
                return []
            
            # First row is header
            columns = rows_data[0]
            data_rows = rows_data[1:]
            df = pd.DataFrame(data_rows, columns=columns)
            
            logger.info(f"Excel loaded: {len(df)} rows, columns: {list(df.columns)}")
            
            # Validate and map columns
            column_map = self._validate_columns(df)
            if not column_map:
                return []
            
            # Parse rows
            rows = []
            for idx, pd_row in df.iterrows():
                row_number = idx + 2  # Excel row (1-indexed, +1 for header)
                
                role_name = str(pd_row[column_map["role name"]]) if pd.notna(pd_row[column_map["role name"]]) else ""
                alias = str(pd_row[column_map["alias"]]) if pd.notna(pd_row[column_map["alias"]]) else ""
                description = str(pd_row[column_map["role description"]]) if pd.notna(pd_row[column_map["role description"]]) else ""
                
                # Skip completely empty rows
                if not role_name.strip() and not alias.strip() and not description.strip():
                    continue
                
                rows.append(RoleImportRow(
                    row_number=row_number,
                    role_name=role_name,
                    alias=alias if alias else None,
                    description=description if description else None
                ))
            
            logger.info(f"Parsed {len(rows)} data rows from Excel")
            return rows
            
        except Exception as e:
            logger.error(f"Failed to parse Excel file: {e}", exc_info=True)
            self.errors.append({
                "row_number": 0,
                "role_name": "",
                "reason": f"Failed to parse Excel file: {str(e)}"
            })
            return []
    
    def _validate_columns(self, df: pd.DataFrame) -> Optional[Dict[str, str]]:
        """Validate Excel columns and return mapping."""
        df_columns_lower = {str(col).strip().lower(): col for col in df.columns}
        
        column_map = {}
        missing = []
        
        for expected in self.EXPECTED_COLUMNS:
            if expected in df_columns_lower:
                column_map[expected] = df_columns_lower[expected]
            else:
                missing.append(expected)
        
        if missing:
            self.errors.append({
                "row_number": 0,
                "role_name": "",
                "reason": f"Missing required columns: {', '.join(missing)}. "
                          f"Found: {', '.join(df.columns)}"
            })
            return None
        
        return column_map
    
    def _build_conflict_maps(self):
        """Build O(1) lookup maps from existing roles."""
        roles = self.db.query(Role).filter(Role.deleted_at.is_(None)).all()
        
        for role in roles:
            # Map role name
            name_norm = normalize_key(role.role_name)
            self.role_name_map[name_norm] = role.role_id
            
            # Map alias tokens
            if role.role_alias:
                for token in self._parse_alias_tokens(role.role_alias):
                    token_norm = normalize_key(token)
                    if token_norm:
                        self.alias_token_map[token_norm] = role.role_id
        
        logger.info(f"Built conflict maps: {len(self.role_name_map)} names, {len(self.alias_token_map)} alias tokens")
    
    def _parse_alias_tokens(self, alias: Optional[str]) -> List[str]:
        """Parse comma-separated alias string into individual tokens."""
        if not alias:
            return []
        return [t.strip() for t in alias.split(",") if t.strip()]
    
    def _check_conflicts(
        self,
        role_name_norm: str,
        alias_tokens_norm: List[str],
        batch_names: Set[str],
        batch_aliases: Set[str]
    ) -> Optional[str]:
        """
        Check for conflicts between new role and existing roles.
        
        Returns conflict reason or None if no conflict.
        """
        # Check role_name vs existing role_names
        if role_name_norm in self.role_name_map:
            return "Duplicate: Role name already exists"
        
        # Check role_name vs batch names
        if role_name_norm in batch_names:
            return "Duplicate: Role name duplicated within import file"
        
        # Check role_name vs existing alias tokens
        if role_name_norm in self.alias_token_map:
            return "Duplicate: Role name matches an existing role's alias"
        
        # Check role_name vs batch aliases
        if role_name_norm in batch_aliases:
            return "Duplicate: Role name matches an alias in earlier import row"
        
        # Check each alias token
        for token_norm in alias_tokens_norm:
            if not token_norm:
                continue
            
            # Alias vs existing role_names
            if token_norm in self.role_name_map:
                return f"Duplicate: Alias matches an existing role name"
            
            # Alias vs batch names
            if token_norm in batch_names:
                return f"Duplicate: Alias matches a role name in earlier import row"
            
            # Alias vs existing alias tokens
            if token_norm in self.alias_token_map:
                return f"Duplicate: Alias already exists for another role"
            
            # Alias vs batch aliases
            if token_norm in batch_aliases:
                return f"Duplicate: Alias duplicated within import file"
        
        return None


def import_roles_from_excel(db: Session, file_content: bytes, created_by: str = "system") -> ImportResult:
    """
    Convenience function to import roles from Excel.
    
    Args:
        db: Database session
        file_content: Excel file bytes
        created_by: User performing import
        
    Returns:
        ImportResult with success/failure counts and details
    """
    service = RoleImportService(db)
    return service.import_roles(file_content, created_by)
