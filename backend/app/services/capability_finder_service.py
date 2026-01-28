"""
Service layer for Capability Finder (Advanced Query) feature.
Handles business logic for fetching typeahead/autocomplete data.
"""
from typing import List, Optional
from sqlalchemy.orm import Session
from sqlalchemy import distinct, func, and_, or_
from app.models.skill import Skill
from app.models.role import Role
from app.models.employee import Employee
from app.models.employee_skill import EmployeeSkill
from app.models.proficiency import ProficiencyLevel
from app.models.sub_segment import SubSegment
from app.models.team import Team
from app.schemas.capability_finder import EmployeeSearchResult, SkillInfo


class CapabilityFinderService:
    """Service for capability finder data operations."""
    
    @staticmethod
    def get_all_skills(db: Session) -> List[str]:
        """
        Get all distinct skill names sorted alphabetically.
        
        Args:
            db: Database session
            
        Returns:
            List of skill names sorted A-Z
        """
        skills = db.query(Skill.skill_name)\
            .distinct()\
            .order_by(Skill.skill_name)\
            .all()
        
        return [skill[0] for skill in skills]
    
    @staticmethod
    def get_all_roles(db: Session) -> List[str]:
        """
        Get all distinct role names sorted alphabetically.
        
        Args:
            db: Database session
            
        Returns:
            List of role names sorted A-Z
        """
        roles = db.query(Role.role_name)\
            .distinct()\
            .order_by(Role.role_name)\
            .all()
        
        return [role[0] for role in roles]
    
    @staticmethod
    def search_matching_talent(
        db: Session,
        skills: List[str],
        sub_segment_id: Optional[int] = None,
        team_id: Optional[int] = None,
        role: Optional[str] = None,
        min_proficiency: int = 0,
        min_experience_years: int = 0
    ) -> List[EmployeeSearchResult]:
        """
        Search for employees matching specified criteria.
        
        Args:
            db: Database session
            skills: List of required skill names (AND logic - must have ALL)
            sub_segment_id: Optional sub-segment filter
            team_id: Optional team filter
            role: Optional role name filter
            min_proficiency: Minimum proficiency level (0-5)
            min_experience_years: Minimum years of experience
            
        Returns:
            List of matching employees with their top 3 skills
        """
        # Build base query for employees who match ALL required skills
        query = db.query(Employee).join(EmployeeSkill).join(Skill)
        
        # Apply filters
        filters = []
          # Skills filter with AND logic
        if skills and len(skills) > 0:
            # Find skill IDs for the required skill names
            skill_ids = db.query(Skill.skill_id)\
                .filter(Skill.skill_name.in_(skills))\
                .all()
            skill_ids = [s[0] for s in skill_ids]
            
            if skill_ids:
                # Subquery to find employees who have ALL required skills
                # with minimum proficiency and experience
                skill_subquery = db.query(EmployeeSkill.employee_id)\
                    .filter(
                        EmployeeSkill.skill_id.in_(skill_ids),
                        EmployeeSkill.proficiency_level_id >= min_proficiency,
                        EmployeeSkill.years_experience >= min_experience_years
                    )\
                    .group_by(EmployeeSkill.employee_id)\
                    .having(func.count(distinct(EmployeeSkill.skill_id)) == len(skill_ids))
                
                filters.append(Employee.employee_id.in_(skill_subquery))
        
        # Organization filters
        if sub_segment_id:
            filters.append(Employee.sub_segment_id == sub_segment_id)
        
        if team_id:
            filters.append(Employee.team_id == team_id)
        
        if role:
            query = query.join(Role, Employee.role_id == Role.role_id)
            filters.append(Role.role_name == role)
        
        # Apply all filters
        if filters:
            query = query.filter(and_(*filters))
        
        # Get distinct employees
        employees = query.distinct().all()
        
        # Build results with top 3 skills for each employee
        results = []
        for employee in employees:            # Get top 3 skills ordered by proficiency DESC, last_used DESC, skill_name ASC
            top_skills_query = db.query(
                Skill.skill_name,
                EmployeeSkill.proficiency_level_id
            )\
                .join(EmployeeSkill, EmployeeSkill.skill_id == Skill.skill_id)\
                .filter(EmployeeSkill.employee_id == employee.employee_id)\
                .order_by(
                    EmployeeSkill.proficiency_level_id.desc(),
                    EmployeeSkill.last_used.desc(),
                    Skill.skill_name.asc()
                )\
                .limit(3)\
                .all()
            
            top_skills = [
                SkillInfo(name=skill_name, proficiency=proficiency)
                for skill_name, proficiency in top_skills_query            ]
            
            # Get organization info
            sub_segment_name = employee.sub_segment.sub_segment_name if employee.sub_segment else ""
            team_name = employee.team.team_name if employee.team else ""
            role_name = employee.role.role_name if employee.role else ""
            
            results.append(EmployeeSearchResult(
                employee_id=employee.employee_id,
                employee_name=employee.full_name,
                sub_segment=sub_segment_name,
                team=team_name,
                role=role_name,
                top_skills=top_skills
            ))
        
        return results
    
    @staticmethod
    def export_matching_talent_to_excel(
        db: Session,
        mode: str,
        skills: List[str],
        sub_segment_id: Optional[int] = None,
        team_id: Optional[int] = None,
        role: Optional[str] = None,
        min_proficiency: int = 0,
        min_experience_years: int = 0,
        selected_employee_ids: List[int] = None
    ):
        """
        Export matching talent to Excel format with all skills consolidated per employee.
        
        Args:
            db: Database session
            mode: 'all' or 'selected'
            skills: List of required skill names (AND logic)
            sub_segment_id: Optional sub-segment filter
            team_id: Optional team filter
            role: Optional role name filter
            min_proficiency: Minimum proficiency level (0-5)
            min_experience_years: Minimum years of experience
            selected_employee_ids: List of employee IDs for mode='selected'
            
        Returns:
            BytesIO object containing Excel file
        """
        import logging
        from io import BytesIO
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        from app.models.project import Project
        from sqlalchemy.orm import joinedload
        
        logger = logging.getLogger(__name__)
        logger.info(f"Export started - mode: {mode}, skills_filter: {skills}, selected_count: {len(selected_employee_ids) if selected_employee_ids else 0}")
        
        # Determine which employee IDs to export
        if mode == 'selected':
            if not selected_employee_ids:
                raise ValueError("selected_employee_ids cannot be empty when mode is 'selected'")
            employee_ids_to_export = set(selected_employee_ids)
            logger.info(f"Exporting {len(employee_ids_to_export)} selected employees")
        else:  # mode == 'all'
            # Get employee IDs from search results (reuse existing search logic)
            search_results = CapabilityFinderService.search_matching_talent(
                db=db,
                skills=skills,
                sub_segment_id=sub_segment_id,
                team_id=team_id,
                role=role,
                min_proficiency=min_proficiency,
                min_experience_years=min_experience_years
            )
            employee_ids_to_export = set(result.employee_id for result in search_results)
            logger.info(f"Exporting all {len(employee_ids_to_export)} matching employees")
        
        if not employee_ids_to_export:
            logger.warning("No employees to export")
            # Create empty workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Matching Talent"
            ws.cell(row=1, column=1, value="No matching employees found")
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        
        # Fetch full employee data with all relationships
        employees = db.query(Employee)\
            .options(
                joinedload(Employee.sub_segment),
                joinedload(Employee.project),
                joinedload(Employee.team),
                joinedload(Employee.role)
            )\
            .filter(Employee.employee_id.in_(employee_ids_to_export))\
            .all()
        
        logger.info(f"Retrieved {len(employees)} employees with full data")
        
        # Build export rows
        export_rows = []
        for employee in employees:
            # Fetch ALL skills for this employee, sorted by proficiency DESC
            employee_skills = db.query(EmployeeSkill, Skill, ProficiencyLevel)\
                .join(Skill, EmployeeSkill.skill_id == Skill.skill_id)\
                .join(ProficiencyLevel, EmployeeSkill.proficiency_level_id == ProficiencyLevel.proficiency_level_id)\
                .filter(EmployeeSkill.employee_id == employee.employee_id)\
                .order_by(EmployeeSkill.proficiency_level_id.desc(), Skill.skill_name.asc())\
                .all()
            
            # Build consolidated skills string (multiline in single cell)
            skills_parts = []
            for emp_skill, skill, prof_level in employee_skills:
                # Start with skill name and proficiency
                prof_text = prof_level.level_name
                # Strip numeric prefix (e.g., "3 - Advanced" -> "Advanced")
                if ' - ' in prof_text:
                    prof_text = prof_text.split(' - ', 1)[1]
                
                skill_parts = [skill.skill_name, f"({prof_text}"]
                
                # Add years of experience if present
                if emp_skill.years_experience is not None and emp_skill.years_experience > 0:
                    skill_parts.append(f", {emp_skill.years_experience}yrs")
                
                # Add last used date if present
                if emp_skill.last_used:
                    last_used_str = emp_skill.last_used.strftime("%Y-%m")
                    skill_parts.append(f", LastUsed: {last_used_str}")
                
                # Add certifications if present
                if emp_skill.certification and emp_skill.certification.strip():
                    # Handle multiple certifications separated by comma or other delimiters
                    certs = emp_skill.certification.replace(',', '|').replace(';', '|')
                    skill_parts.append(f", Certs: {certs}")
                
                skill_parts.append(")")
                skills_parts.append(''.join(skill_parts))
            
            # Join all skills with semicolon + newline
            skills_text = ";\n".join(skills_parts) + (";" if skills_parts else "")
            
            export_rows.append({
                'employee_name': employee.full_name,
                'zid': employee.zid,
                'sub_segment': employee.sub_segment.sub_segment_name if employee.sub_segment else "",
                'project': employee.project.project_name if employee.project else "",
                'team': employee.team.team_name if employee.team else "",
                'role': employee.role.role_name if employee.role else "",
                'skills': skills_text
            })
        
        logger.info(f"Built {len(export_rows)} export rows")
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Matching Talent"
        
        # Define headers in exact order
        headers = ["Employee Name", "ZID", "Sub-segment", "Project Name", "Team Name", "Role", "Skills"]
          # Style for header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="left", vertical="center")
        
        # Alignment for data cells (vertically centered)
        data_alignment = Alignment(vertical="center")
        data_alignment_wrapped = Alignment(wrap_text=True, vertical="center")
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        ws.row_dimensions[1].height = 35
        # Write data rows with vertical centering
        for row_idx, row_data in enumerate(export_rows, start=2):
            cell = ws.cell(row=row_idx, column=1, value=row_data['employee_name'])
            cell.alignment = data_alignment
            
            cell = ws.cell(row=row_idx, column=2, value=row_data['zid'])
            cell.alignment = data_alignment
            
            cell = ws.cell(row=row_idx, column=3, value=row_data['sub_segment'])
            cell.alignment = data_alignment
            
            cell = ws.cell(row=row_idx, column=4, value=row_data['project'])
            cell.alignment = data_alignment
            
            cell = ws.cell(row=row_idx, column=5, value=row_data['team'])
            cell.alignment = data_alignment
            
            cell = ws.cell(row=row_idx, column=6, value=row_data['role'])
            cell.alignment = data_alignment
            
            # Skills column with wrap text and vertical centering
            skills_cell = ws.cell(row=row_idx, column=7, value=row_data['skills'])
            skills_cell.alignment = data_alignment_wrapped
        
        # Set column widths explicitly
        # Columns A-F (Employee Name, ZID, Sub-segment, Project Name, Team Name, Role): width 20
        for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col_letter].width = 20
        
        # Column G (Skills): width 75 for multiline content
        ws.column_dimensions['G'].width = 75
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        logger.info(f"Excel file created successfully with {len(export_rows)} rows")
        return output
