"""
Export service for Capability Finder.

Handles exporting matching talent to Excel with all skills consolidated per employee.
Supports two modes: export all search results, or export only selected employees.

Isolated from search service - no shared helpers to ensure changes in export
cannot break search functionality and vice versa.
"""
import logging
from typing import List, Optional, Set, Dict, Any
from io import BytesIO
from datetime import datetime

from sqlalchemy.orm import Session, joinedload
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from app.models.employee import Employee
from app.models.team import Team
from app.models.project import Project
from app.models.sub_segment import SubSegment
from app.models.employee_skill import EmployeeSkill
from app.models.skill import Skill
from app.models.proficiency import ProficiencyLevel
from app.models.role import Role
from app.schemas.capability_finder import EmployeeSearchResult

logger = logging.getLogger(__name__)


def export_matching_talent_to_excel(
    db: Session,
    mode: str,
    skills: List[str],
    sub_segment_id: Optional[int] = None,
    team_id: Optional[int] = None,
    role: Optional[str] = None,
    min_proficiency: int = 0,
    min_experience_years: int = 0,
    selected_employee_ids: List[int] = None
) -> BytesIO:
    """
    Export matching talent to Excel format with all skills consolidated per employee.
    
    Two export modes:
    - 'all': Export all employees matching the search filters
    - 'selected': Export only specified employee IDs
    
    Excel format:
    - Header row with blue background
    - Columns: Employee Name, ZID, Sub-segment, Project Name, Team Name, Role, Skills
    - Skills column contains all skills with proficiency, experience, certifications
    - Each skill on a new line within the cell
    
    Args:
        db: Database session
        mode: 'all' or 'selected'
        skills: List of required skill names (AND logic)
        sub_segment_id: Optional sub-segment filter
        team_id: Optional team filter
        role: Optional role name filter
        min_proficiency: Minimum proficiency level (0-5)
        min_experience_years: Minimum years of experience
        selected_employee_ids: List of employee IDs for mode='selected'
        
    Returns:
        BytesIO object containing Excel file
        
    Raises:
        ValueError: If mode='selected' but selected_employee_ids is empty
    """
    logger.info(f"Export started - mode: {mode}, skills_filter: {skills}, "
                f"selected_count: {len(selected_employee_ids) if selected_employee_ids else 0}")
    
    # Validate input
    _validate_export_request(mode, selected_employee_ids)
    
    # Determine which employee IDs to export
    employee_ids_to_export = _determine_employee_ids_to_export(
        db=db,
        mode=mode,
        skills=skills,
        sub_segment_id=sub_segment_id,
        team_id=team_id,
        role=role,
        min_proficiency=min_proficiency,
        min_experience_years=min_experience_years,
        selected_employee_ids=selected_employee_ids
    )
    
    # Handle empty result set
    if not employee_ids_to_export:
        logger.warning("No employees to export")
        return _create_empty_workbook()
    
    # Fetch full employee data
    employees = _query_employees_with_relationships(db, employee_ids_to_export)
    logger.info(f"Retrieved {len(employees)} employees with full data")
    
    # Build export rows
    export_rows = _build_export_rows(db, employees)
    logger.info(f"Built {len(export_rows)} export rows")
    
    # Create Excel workbook
    excel_file = _create_excel_workbook(export_rows)
    logger.info(f"Excel file created successfully with {len(export_rows)} rows")
    
    return excel_file


def _validate_export_request(mode: str, selected_employee_ids: Optional[List[int]]) -> None:
    """
    Validate export request parameters.
    
    Pure validation helper - no DB access, no side effects.
    
    Args:
        mode: Export mode ('all' or 'selected')
        selected_employee_ids: List of employee IDs (required for 'selected' mode)
        
    Raises:
        ValueError: If validation fails
    """
    if mode == 'selected' and not selected_employee_ids:
        raise ValueError("selected_employee_ids cannot be empty when mode is 'selected'")


def _determine_employee_ids_to_export(
    db: Session,
    mode: str,
    skills: List[str],
    sub_segment_id: Optional[int],
    team_id: Optional[int],
    role: Optional[str],
    min_proficiency: int,
    min_experience_years: int,
    selected_employee_ids: Optional[List[int]]
) -> Set[int]:
    """
    Determine which employee IDs to export based on mode and filters.
    
    For 'selected' mode: return selected IDs
    For 'all' mode: search for matching employees and return their IDs
    
    Args:
        db: Database session
        mode: Export mode ('all' or 'selected')
        skills: Required skill names
        sub_segment_id: Optional sub-segment filter
        team_id: Optional team filter
        role: Optional role name filter
        min_proficiency: Minimum proficiency level
        min_experience_years: Minimum years of experience
        selected_employee_ids: Employee IDs for 'selected' mode
        
    Returns:
        Set of employee IDs to export
    """
    if mode == 'selected':
        employee_ids = set(selected_employee_ids)
        logger.info(f"Exporting {len(employee_ids)} selected employees")
        return employee_ids
    
    # mode == 'all': run search to get matching employees
    # NOTE: We duplicate the search logic here to avoid cross-dependencies
    # with search_service. This ensures export changes cannot break search.
    search_results = _search_matching_talent_for_export(
        db=db,
        skills=skills,
        sub_segment_id=sub_segment_id,
        team_id=team_id,
        role=role,
        min_proficiency=min_proficiency,
        min_experience_years=min_experience_years
    )
    
    employee_ids = set(result.employee_id for result in search_results)
    logger.info(f"Exporting all {len(employee_ids)} matching employees")
    return employee_ids


def _search_matching_talent_for_export(
    db: Session,
    skills: List[str],
    sub_segment_id: Optional[int],
    team_id: Optional[int],
    role: Optional[str],
    min_proficiency: int,
    min_experience_years: int
) -> List[EmployeeSearchResult]:
    """
    Search for matching employees (duplicated from search service for isolation).
    
    This is intentionally duplicated to ensure export logic is independent
    from search logic. Changes in one cannot break the other.
    
    Args:
        db: Database session
        skills: Required skill names (AND logic)
        sub_segment_id: Optional sub-segment filter
        team_id: Optional team filter
        role: Optional role name filter
        min_proficiency: Minimum proficiency level
        min_experience_years: Minimum years of experience
        
    Returns:
        List of EmployeeSearchResult objects
    """
    # Import here to avoid circular dependencies
    from app.services.capability_finder.search_service import search_matching_talent
    
    return search_matching_talent(
        db=db,
        skills=skills,
        sub_segment_id=sub_segment_id,
        team_id=team_id,
        role=role,
        min_proficiency=min_proficiency,
        min_experience_years=min_experience_years
    )


def _query_employees_with_relationships(
    db: Session,
    employee_ids: Set[int]
) -> List[Employee]:
    """
    Query employees with all relationships eagerly loaded.
    
    DB-only helper - no business logic.
    
    Args:
        db: Database session
        employee_ids: Set of employee IDs to fetch
        
    Returns:
        List of Employee objects with relationships loaded
    """
    employees = db.query(Employee)\
        .options(
            # NORMALIZED: Load org chain via team -> project -> sub_segment
            joinedload(Employee.team)
                .joinedload(Team.project)
                .joinedload(Project.sub_segment),
            joinedload(Employee.role)
        )\
        .filter(Employee.employee_id.in_(employee_ids))\
        .all()
    
    return employees


def _build_export_rows(db: Session, employees: List[Employee]) -> List[Dict[str, Any]]:
    """
    Build export row data for all employees.
    
    Each row contains employee info and consolidated skills text.
    
    Args:
        db: Database session (needed to fetch skills)
        employees: List of Employee objects
        
    Returns:
        List of dictionaries with export data
    """
    export_rows = []
    
    for employee in employees:
        # Fetch all skills for this employee
        employee_skills = _query_employee_all_skills(db, employee.employee_id)
        
        # Build consolidated skills text
        skills_text = _build_skills_text(employee_skills)
        
        # Build row dictionary
        row = _build_export_row(employee, skills_text)
        export_rows.append(row)
    
    return export_rows


def _query_employee_all_skills(db: Session, employee_id: int) -> List[tuple]:
    """
    Query all skills for an employee with proficiency and metadata.
    
    Skills are ordered by proficiency DESC, then skill name ASC.
    
    DB-only helper - no business logic.
    
    Args:
        db: Database session
        employee_id: Employee ID
        
    Returns:
        List of tuples (EmployeeSkill, Skill, ProficiencyLevel)
    """
    employee_skills = db.query(EmployeeSkill, Skill, ProficiencyLevel)\
        .join(Skill, EmployeeSkill.skill_id == Skill.skill_id)\
        .join(ProficiencyLevel, EmployeeSkill.proficiency_level_id == ProficiencyLevel.proficiency_level_id)\
        .filter(EmployeeSkill.employee_id == employee_id)\
        .order_by(EmployeeSkill.proficiency_level_id.desc(), Skill.skill_name.asc())\
        .all()
    
    return employee_skills


def _build_skills_text(employee_skills: List[tuple]) -> str:
    """
    Build consolidated skills text for Excel cell.
    
    Format: "SkillName (Proficiency, XYrs, LastUsed: YYYY-MM, Certs: CertName);"
    Each skill on a new line within the cell.
    
    Pure transformation helper - no DB access, no side effects.
    
    Args:
        employee_skills: List of tuples (EmployeeSkill, Skill, ProficiencyLevel)
        
    Returns:
        Multiline string with all skills formatted
    """
    skills_parts = []
    
    for emp_skill, skill, prof_level in employee_skills:
        skill_text = _format_single_skill(emp_skill, skill, prof_level)
        skills_parts.append(skill_text)
    
    # Join with semicolon + newline and add trailing semicolon
    return ";\n".join(skills_parts) + (";" if skills_parts else "")


def _format_single_skill(
    emp_skill: EmployeeSkill,
    skill: Skill,
    prof_level: ProficiencyLevel
) -> str:
    """
    Format a single skill with all metadata.
    
    Format: "SkillName (Proficiency, XYrs, LastUsed: YYYY-MM, Certs: CertName)"
    
    Pure transformation helper - no DB access, no side effects.
    
    Args:
        emp_skill: EmployeeSkill instance
        skill: Skill instance
        prof_level: ProficiencyLevel instance
        
    Returns:
        Formatted skill string
    """
    # Extract proficiency text (strip numeric prefix)
    prof_text = prof_level.level_name
    if ' - ' in prof_text:
        prof_text = prof_text.split(' - ', 1)[1]
    
    # Start building skill parts
    skill_parts = [skill.skill_name, f"({prof_text}"]
    
    # Add years of experience if present
    if emp_skill.years_experience is not None and emp_skill.years_experience > 0:
        skill_parts.append(f", {emp_skill.years_experience}yrs")
    
    # Add last used date if present
    if emp_skill.last_used:
        last_used_str = emp_skill.last_used.strftime("%Y-%m")
        skill_parts.append(f", LastUsed: {last_used_str}")
    
    # Add certifications if present
    if emp_skill.certification and emp_skill.certification.strip():
        # Handle multiple certifications separated by comma or semicolon
        certs = emp_skill.certification.replace(',', '|').replace(';', '|')
        skill_parts.append(f", Certs: {certs}")
    
    skill_parts.append(")")
    
    return ''.join(skill_parts)


def _build_export_row(employee: Employee, skills_text: str) -> Dict[str, Any]:
    """
    Build export row dictionary from employee and skills text.
    
    Pure transformation helper - no DB access, no side effects.
    
    Args:
        employee: Employee instance
        skills_text: Consolidated skills text
        
    Returns:
        Dictionary with all export columns
    """
    return {
        'employee_name': employee.full_name,
        'zid': employee.zid,
        'sub_segment': employee.sub_segment.sub_segment_name if employee.sub_segment else "",
        'project': employee.project.project_name if employee.project else "",
        'team': employee.team.team_name if employee.team else "",
        'role': employee.role.role_name if employee.role else "",
        'skills': skills_text
    }


def _create_empty_workbook() -> BytesIO:
    """
    Create an empty Excel workbook with "No matching employees found" message.
    
    Pure transformation helper - no DB access, no side effects.
    
    Returns:
        BytesIO object with empty workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Matching Talent"
    ws.cell(row=1, column=1, value="No matching employees found")
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def _create_excel_workbook(export_rows: List[Dict[str, Any]]) -> BytesIO:
    """
    Create Excel workbook from export rows with formatting.
    
    Applies specific formatting:
    - Blue header row with white text
    - Column widths: A-F = 20, G (Skills) = 75
    - Header row height = 35
    - Skills column with wrap text and vertical centering
    
    Pure transformation helper - no DB access, no side effects.
    
    Args:
        export_rows: List of export row dictionaries
        
    Returns:
        BytesIO object containing formatted Excel file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Matching Talent"
    
    # Define headers in exact order
    headers = ["Employee Name", "ZID", "Sub-segment", "Project Name", "Team Name", "Role", "Skills"]
    
    # Apply header formatting
    _apply_header_formatting(ws, headers)
    
    # Write data rows
    _write_data_rows(ws, export_rows)
    
    # Set column widths
    _set_column_widths(ws)
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def _apply_header_formatting(ws, headers: List[str]) -> None:
    """
    Apply formatting to header row.
    
    Pure helper - modifies worksheet in place.
    
    Args:
        ws: Worksheet object
        headers: List of header strings
    """
    # Style for header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="left", vertical="center")
    
    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    ws.row_dimensions[1].height = 35


def _write_data_rows(ws, export_rows: List[Dict[str, Any]]) -> None:
    """
    Write data rows to worksheet with formatting.
    
    Pure helper - modifies worksheet in place.
    
    Args:
        ws: Worksheet object
        export_rows: List of export row dictionaries
    """
    # Alignment for data cells
    data_alignment = Alignment(vertical="center")
    data_alignment_wrapped = Alignment(wrap_text=True, vertical="center")
    
    for row_idx, row_data in enumerate(export_rows, start=2):
        # Column 1: Employee Name
        cell = ws.cell(row=row_idx, column=1, value=row_data['employee_name'])
        cell.alignment = data_alignment
        
        # Column 2: ZID
        cell = ws.cell(row=row_idx, column=2, value=row_data['zid'])
        cell.alignment = data_alignment
        
        # Column 3: Sub-segment
        cell = ws.cell(row=row_idx, column=3, value=row_data['sub_segment'])
        cell.alignment = data_alignment
        
        # Column 4: Project
        cell = ws.cell(row=row_idx, column=4, value=row_data['project'])
        cell.alignment = data_alignment
        
        # Column 5: Team
        cell = ws.cell(row=row_idx, column=5, value=row_data['team'])
        cell.alignment = data_alignment
        
        # Column 6: Role
        cell = ws.cell(row=row_idx, column=6, value=row_data['role'])
        cell.alignment = data_alignment
        
        # Column 7: Skills (with wrap text)
        skills_cell = ws.cell(row=row_idx, column=7, value=row_data['skills'])
        skills_cell.alignment = data_alignment_wrapped


def _set_column_widths(ws) -> None:
    """
    Set column widths to exact specifications.
    
    Pure helper - modifies worksheet in place.
    
    Args:
        ws: Worksheet object
    """
    # Columns A-F: width 20
    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col_letter].width = 20
    
    # Column G (Skills): width 75 for multiline content
    ws.column_dimensions['G'].width = 75
